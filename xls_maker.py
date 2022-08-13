import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
                        PatternFill, Border, Side,
                        Alignment, Font, GradientFill
                        )
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule


def create_xls_fie(lst, filename="Order.xls"):
    df = pd.DataFrame(lst)

    df = df.reindex(columns=["Код", "Люмн. Код", "Наименование","ISBN",
                             "Издательство", "Год издания", "Количество",
                             "Интер. Остаток", "Интер. Цена", "Интер. Цена со скидкой", "Заказ",
                             "Люмн. Остаток", "Люмн. Цена", "Люмн. Цена со скидкой", "Заказ (количество)",
                             "Проверка"])

    df.to_excel(filename, index=False)
    return filename


def arrange_xls_file(filename="C:\\Users\\User1\\Downloads\\order2.xlsx"):
    wb = load_workbook(filename=filename)
    sheet = wb.active
    rows_amount = len(sheet["A"])

    sheet["B1"].value="Код"
    sheet["V1"].value = 0

    amount_font = Font(size=13, bold=True, )
    col_amount = sheet["G"]
    for elem in col_amount:
        if type(elem.value) == int:
            elem.font = amount_font
            elem.alignment = Alignment(horizontal="center")

    inter_cols = ["I", "J", "H"]
    inter_cost_cols = ["I", "J"]
    fill_cols(inter_cols, inter_cost_cols, "00FFFF00", sheet)

    lumn_cols = ["M", "N", "L"]
    lumn_cost_cols = ["M", "N"]
    fill_cols(lumn_cols, lumn_cost_cols, "0000FF00", sheet)

    for index, elem in enumerate(sheet["P"]):
        if elem.value == "Проверка":
            continue
        else:
            elem.value = f"=K{index+1}+O{index+1}-G{index+1}"

    sheet.conditional_formatting.add(f"P2:P{rows_amount}", CellIsRule(operator='greaterThan', formula=['V$1'],
                                                            fill=PatternFill(fill_type='solid', fgColor='00FF0000')))

    wb.freeze_panes = "Q2"

    wb.save(filename)


def fill_cols(cols, cost_cols, color, sheet):
    border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="medium"),
                    bottom=Side(border_style="thin"), diagonal=Side(border_style="thin"))
    for col in cols:
        column = sheet[col]
        if col in cost_cols:
            for elem in column:
                elem.alignment = Alignment(horizontal="center")
                elem.fill = PatternFill('solid', fgColor=color)
                elem.border = border
        else:
            for elem in column:
                elem.fill = PatternFill('solid', fgColor=color)
                elem.border = border


# arrange_xls_file()
